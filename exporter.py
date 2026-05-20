"""Export financial data from SQLite to a formatted Excel workbook."""

import logging
from datetime import datetime
from typing import Optional

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

from database import Database

logger = logging.getLogger(__name__)

# ── Style constants ───────────────────────────────────────────────────────────

_HEADER_FILL   = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT   = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
_ALT_FILL      = PatternFill("solid", fgColor="D6E4F0")
_NUMBER_FORMAT = '#,##0.00'
_THIN          = Side(style="thin", color="BFBFBF")
_BORDER        = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_COLUMNS = [
    ("Company Name",    "company_name",  40),
    ("Year",            "year",          8),
    ("Revenue",         "revenue",       18),
    ("Net Profit",      "net_profit",    18),
    ("Total Assets",    "assets",        18),
    ("Total Liab.",     "liabilities",   18),
    ("Equity",          "equity",        18),
    ("Scraped At",      "scraped_at",    22),
]


# ── public API ────────────────────────────────────────────────────────────────

def export_to_excel(db: Database, output_path: str) -> str:
    """
    Write all financial_data rows to *output_path* as a formatted .xlsx file.
    Returns the resolved output path.
    """
    records = db.get_all_financial_data()
    if not records:
        logger.warning("No financial data found — empty workbook will be written.")

    wb = Workbook()
    _build_data_sheet(wb, records)
    _build_summary_sheet(wb, records)

    wb.save(output_path)
    logger.info("Excel report saved → %s  (%d rows)", output_path, len(records))
    return output_path


# ── sheet builders ────────────────────────────────────────────────────────────

def _build_data_sheet(wb: Workbook, records: list[dict]):
    ws = wb.active
    ws.title = "Financial Data"
    ws.freeze_panes = "A2"

    # Header row
    for col_idx, (header, _, width) in enumerate(_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font   = _HEADER_FONT
        cell.fill   = _HEADER_FILL
        cell.border = _BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 30

    # Data rows
    for row_idx, record in enumerate(records, start=2):
        fill = _ALT_FILL if row_idx % 2 == 0 else None
        for col_idx, (_, key, _) in enumerate(_COLUMNS, start=1):
            value = record.get(key)
            cell  = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = _BORDER
            if fill:
                cell.fill = fill

            # Right-align numbers
            if key in ("revenue", "net_profit", "assets", "liabilities", "equity"):
                cell.alignment  = Alignment(horizontal="right")
                cell.number_format = _NUMBER_FORMAT
            elif key == "year":
                cell.alignment = Alignment(horizontal="center")
            elif key == "scraped_at":
                cell.alignment  = Alignment(horizontal="center")
                cell.number_format = "YYYY-MM-DD HH:MM:SS"
            else:
                cell.alignment = Alignment(horizontal="left", wrap_text=True)

    # Auto-filter
    ws.auto_filter.ref = ws.dimensions


def _build_summary_sheet(wb: Workbook, records: list[dict]):
    """Pivot: one row per company, latest-year values, plus simple ratios."""
    ws = wb.create_sheet("Summary")
    ws.freeze_panes = "A2"

    summary_cols = [
        ("Company Name",         40),
        ("Latest Year",          12),
        ("Revenue",              18),
        ("Net Profit",           18),
        ("Profit Margin %",      18),
        ("Total Assets",         18),
        ("Total Liab.",          18),
        ("Equity",               18),
        ("Debt/Equity Ratio",    18),
    ]

    for col_idx, (header, width) in enumerate(summary_cols, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font   = _HEADER_FONT
        cell.fill   = _HEADER_FILL
        cell.border = _BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 30

    # Build latest record per company
    latest: dict[str, dict] = {}
    for r in records:
        name = r["company_name"]
        if name not in latest or (r["year"] or 0) > (latest[name]["year"] or 0):
            latest[name] = r

    for row_idx, (name, r) in enumerate(sorted(latest.items()), start=2):
        rev    = r.get("revenue")
        profit = r.get("net_profit")
        assets = r.get("assets")
        liab   = r.get("liabilities")
        equity = r.get("equity")

        margin = (profit / rev * 100) if rev else None
        de_ratio = (liab / equity) if equity else None

        fill = _ALT_FILL if row_idx % 2 == 0 else None
        values = [name, r.get("year"), rev, profit, margin, assets, liab, equity, de_ratio]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = _BORDER
            if fill:
                cell.fill = fill
            if col_idx == 1:
                cell.alignment = Alignment(horizontal="left", wrap_text=True)
            elif col_idx == 2:
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = '#,##0.00'

    ws.auto_filter.ref = ws.dimensions
